
def add_module():
    # from faulthandler import disable
    import tkinter as tk
    from openpyxl import load_workbook
    import functions.收據編號產生器 as 收據編號產生器
    from datetime import datetime
    import 數字轉換器 as converter
    from 輸出列表機 import printing
    from 收據系統 import menu
    import 編碼器
    import shutil
    window = tk.Tk()
    window.title("中 研 福 德 宮")
    window.geometry("650x450+300+150")
    window.iconbitmap("image\Puyma32x32.ico")
    window.resizable(0,0) #限制視窗縮放 0= true ,1 = fals 
    window.config(bg="skyblue")
    window.attributes("-topmost",1) #置頂 0= true ,1 = fals

    def now_user(): #找出目前使用者、會計及負責人
        last_user = ""
        time2 = ""
        pic=""
        accountant=""
        with open("./database/password.txt", mode="r",encoding="UTF-8") as f:
                    for line in f.readlines():
                        time1 = (line.split(','))[4]
                        if time1 > time2:
                            time2 = time1
                            last_user = (line.split(','))[2] 
                        if (line.split(','))[1] == "負責人":
                            pic = (line.split(','))[2]
                        if (line.split(','))[1] == "會計":
                            accountant = (line.split(','))[2]
                                          
        return last_user,pic,accountant

    # def pic_acct(): #找出會計及負責人
    #     pic=""
    #     accountant=""
    #     with open("./database/password.txt", mode="r",encoding="UTF-8") as f:
    #                 for line in f.readlines():
    #                     time1 = (line.split(','))[4]
    #                     if (line.split(','))[1] == "負責人":
    #                         pic = (line.split(','))[2]
    #                     elif (line.split(','))[1] == "會計":
    #                         accountant = (line.split(','))[2]                 
    #     return pic,accountant


    label_head = tk.Label(window,text="新  增  列  印",font=("標楷體",24,'bold'),
                    relief="ridge",fg='#000',bg="#FFFFCC" ,pady=2,padx=5)      
    add_lab1 = tk.Label(window,text="  收據編號  ",font=("Arial",14,'bold'),
                    fg='#fff',bg='#f50',relief="raised",pady=2,padx=5)                        
    add_lab1_1 = tk.Label(window,text=str(收據編號產生器.new_ser_No),
                    font=("Arial",14,'bold'),relief="raised",pady=2,padx=18,bg="gray72",fg="blue")        
    add_lab2 = tk.Label(window, text="  收款日期  ",font=("Arial",14,'bold'),
                    fg='#fff',bg='#f50', relief="raised",pady=2,padx=5)               
    add_lab2_1 = tk.Label(window, text=str(datetime.today())[0:10],
                    font=("Arial",14,'bold'),relief="raised",pady=2, padx=5,bg="gray72",fg="blue")            
    add_lab3 = tk.Label(window,text="  付  款  人  ",
                    font=("Arial",14,'bold'),fg='#fff',bg='#f50',
                    relief="raised", pady=2, padx=5)       
    add_lab4 = tk.Label(window,text="  地        址  ",
                    font=("Arial",14,'bold'),fg='#fff',bg='#f50',
                    relief="raised", pady=2, padx=5)      
    add_lab5 = tk.Label(window,text="  收款性質  ",
                    font=("Arial",14,'bold'),fg='#fff',bg='#f50',
                    relief="raised", pady=2, padx=5)      
    add_lab6 = tk.Label(window,text="  收款金額  ",
                    font=("Arial",14,'bold'),fg='#fff',bg='#f50',
                    relief="raised", pady=2, padx=5)                 
    add_lab7 = tk.Label(window,text="  經  手  人  ",
                    font=("Arial",14,'bold'),fg='#fff',bg='#f50',
                    relief="raised", pady=2, padx=5)        
    add_lab7_1 = tk.Label(window,text=now_user()[0],font=("Arial",14,'bold'),
                    width=9,relief="raised",pady=2,padx=5,bg="gray72",fg="blue")                                                         
    add_lab8 = tk.Label(window,text="  負  責  人  ",font=("Arial",14,'bold'),
                    fg='#fff', bg='#f50', relief="raised",pady=2,padx=5)   
    add_lab8_1 = tk.Label(window,text=now_user()[1],font=("Arial",14,'bold'),
                    width=9,relief="raised",pady=2, padx=5,bg="gray72",fg="blue") 
    add_lab9 = tk.Label(window,text="  會        計  ",font=("Arial",14,'bold'),
                    fg='#fff', bg='#f50', relief="raised",pady=2,padx=5)   
    add_lab9_1 = tk.Label(window,text=now_user()[2],font=("Arial",14,'bold'),
                    width=9,relief="raised",pady=2, padx=5,bg="gray72",fg="blue")                                           
    add_lab10 = tk.Label(window,text="列印前、請再次確認輸入資料是否正確 !",
                font=("Arial",12,'bold'),fg='#f00',relief='groove', padx=5,pady=2) 
    label_copyright = tk.Label(window,text="Copyright © 2023 Puyuma Workshop ",
                    font=("Arial",8,'bold'),bg="skyblue" ,padx=1,pady=1) 
    def validate(P): 
        if str.isdigit(P) or P == '':
            return True
        else:
            return False

    entry3=tk.Entry(font=('Arial',14,'bold'),insertofftime = 50000,width=35,border=1,bd=5)
    entry4=tk.Entry(font=('Arial',14,'bold'),width=35,border=1,bd=5)
    #####################################################
    def enable():
        entry5.config(state="normal")
    def disable():
        entry5.config(state="disable")
    val = tk.StringVar()  # 建立文字變數
    radio_btn1 = tk.Radiobutton(text='喜添捐獻',font=('Arial',12),bg="skyblue",
                                variable=val, value='喜添捐獻',command=disable)
    radio_btn1.place(x=200, y=235)
    radio_btn1.select()
    radio_btn2 = tk.Radiobutton(text='入會費',font=('Arial',12),bg="skyblue",
                                variable=val, value='入會費',command=disable)
    radio_btn2.place(x=305, y=235)
    radio_btn3 = tk.Radiobutton(text='年度會費',font=('Arial',12),bg="skyblue",
                                variable=val, value='年度會費',command=enable)
    radio_btn3.place(x=400, y=235)
    radio_btn4 = tk.Radiobutton(text='其他',font=('Arial',12),bg="skyblue",
                                variable=val, value='其他',command=enable)
    radio_btn4.place(x=500, y=235)
    entry5=tk.Entry(textvariable=val, font=('Arial',14,'bold'),width=15,border=1,bd=5,state="disable")

    ######################################################
    vcmd = (window.register(validate), '%P') #限制只能輸入數字
    entry6=tk.Entry(window, validate='key', validatecommand=vcmd,font=('Arial',14,'bold'),
                    width=10,border=1,bd=5)
    def clickbton():
        data = add_lab1_1.cget("text")+ ","+ str(datetime.today())[0:10]+ ","+ str(entry3.get())+ ","+ str(entry4.get())+","+str(entry5.get())+","+str(entry6.get())+","+now_user()[0]+","+now_user()[1]+","+now_user()[2]+",,\n"
        with open("./database/record.txt", mode="a+", encoding="utf-8-sig") as f:
            f.write(data)    
        wb = load_workbook("收據範本1210.xlsx")
        ws = wb.active
        # 存根聯
        ws["f14"].value = "NO. " + add_lab1_1.cget("text")
        ws["c4"].value = str(int(str(datetime.today())[0:4])-1911)+"年"+str(datetime.today())[5:7]+"月"+str(datetime.today())[8:10]+"日"
        ws["b5"].value = entry3.get()
        ws["b7"].value = entry4.get()
        ws["c8"].value = entry5.get()
        ws["c9"].value = converter.m_convert(entry6.get())
        ws["b10"].value = now_user()[0]
        ws["f10"].value = now_user()[1]
        ws["d10"].value = now_user()[2]
        # 收執聯
        ws["f33"].value = "NO. " + add_lab7_1.cget("text")
        ws["c23"].value = str(int(str(datetime.today())[0:4])-1911)+"年"+str(datetime.today())[5:7]+"月"+str(datetime.today())[8:10]+"日"
        ws["b24"].value = entry3.get()
        ws["b26"].value = entry4.get()
        ws["c27"].value = entry5.get()
        ws["c28"].value = converter.m_convert(entry6.get())
        ws["b29"].value = now_user()[0]
        ws["f29"].value = now_user()[1]
        ws["d29"].value = now_user()[2]
        # wb.save("中研福德宮"+str(收據編號產生器.new_ser_No)+".xlsx")
        wb.save("收據範本1210.xlsx")
        printing(wb) ###### 輸出列表機 ######
        try: # 資料備份到e:\
            shutil.copy('./database/record.txt','e:./record_backup')
        except:
            pass 
        #add_btn.configure(state=tk.DISABLED)
        add_lab1_1.config(text="0"+str(int(add_lab1_1.cget("text"))+1))
        entry3.delete(0,'end')
        entry4.delete(0,'end')    
        entry5.delete(0,'end')
        entry6.delete(0,'end')
        
    def ending():
        window.destroy()
        menu()

    add_btn = tk.Button(window,text="儲存 與 列印",font=('Arial',14,'bold'),
                    padx=2,pady=2,activeforeground='#f00', command=clickbton)        
    ending_btn = tk.Button(window,text="離  開",font=('Arial',12,'bold'),
                    fg="white",bg="blue",activeforeground='#f00',command=ending)

    label_head.place(x=200,y=5)
    add_lab1.place(x=340, y=65)
    add_lab1_1.place(x=470, y=65)
    add_lab2.place(x=70, y=65)
    add_lab2_1.place(x=200, y=65)
    add_lab3.place(x=70, y=107)
    entry3.place(x=200, y=105)
    add_lab4.place(x=70, y=150)
    entry4.place(x=200, y=147)
    add_lab5.place(x=70, y=193)
    entry5.place(x=200, y=189)
    add_lab6.place(x=70, y=270)
    entry6.place(x=200, y=267)
    add_lab7.place(x=70, y=310)
    add_lab7_1.place(x=200, y=310)
    add_lab8.place(x=340, y=310)
    add_lab8_1.place(x=470, y=310)
    add_lab9.place(x=340, y=270)
    add_lab9_1.place(x=470, y=270)

    add_lab10.place(x=180, y=350)
    add_btn.place(x=260, y=390)
    ending_btn.place(x=40, y=400)
    label_copyright.place(x=430, y=430)
    window.mainloop()

if __name__ == "__main__":
    add_module()