def mod_module():
    import tkinter as tk
    from openpyxl import load_workbook
    import functions.收據編號產生器 as 收據編號產生器
    from datetime import datetime
    import 數字轉換器 as converter
    from 輸出列表機 import printing
    from tkinter import messagebox
    import shutil
    from 收據系統 import menu
    window = tk.Tk()
    window.title("中 研 福 德 宮")
    window.geometry("650x450+300+150")
    window.iconbitmap("Puyma32x32.ico")
    window.resizable(0,0) #限制視窗縮放 0= true ,1 = fals 
    window.config(bg="skyblue")
    window.attributes("-topmost",1) #置頂 0= true ,1 = fals
    label_head = tk.Label(window,text="修  改  列  印",font=("標楷體",24,'bold'),
                    fg='#000',bg="#FFFFCC" ,relief="ridge",pady=2,padx=5)      
    mod_lab1 = tk.Label(window,text="  收據編號  ",font=("Arial",14,'bold'),
                    fg='#fff',bg='#f50',relief="raised",pady=2,padx=5,)           
    mod_lab2 = tk.Label(window, text="  捐款日期  ",font=("Arial",14,'bold'),
                    fg='#fff',bg='#f50', relief="raised",pady=2,padx=5)               
    mod_lab2_1 = tk.Label(window, font=("Arial",14,'bold'),
                    width=9,relief="raised",pady=2, padx=5,bg="gray72",fg="blue")            
    mod_lab3 = tk.Label(window,text="  付  款  人  ",font=("Arial",14,'bold'),
                    fg='#fff',bg='#f50',relief="raised", pady=2, padx=5)   
    mod_lab4 = tk.Label(window,text="  地        址  ",font=("Arial",14,'bold'),
                    fg='#fff',bg='#f50',
                    relief="raised", pady=2, padx=5)      
    mod_lab5 = tk.Label(window,text="  收款性質  ",font=("Arial",14,'bold'),
                    fg='#fff',bg='#f50',relief="raised", pady=2, padx=5)                               
    mod_lab6 = tk.Label(window,text="  付款金額  ",font=("Arial",14,'bold'),
                    fg='#fff',bg='#f50',relief="raised", pady=2, padx=5)                                 
    mod_lab7 = tk.Label(window,text="  經  手  人  ",font=("Arial",14,'bold'),
                    fg='#fff',bg='#f50',relief="raised", pady=2, padx=5)                   
    mod_lab7_1 = tk.Label(window,font=("Arial",14,'bold'),
                    width=9,relief="raised",pady=2,padx=5,bg="gray72",fg="blue")                  
    mod_lab8 = tk.Label(window,text="  負  責  人  ",font=("Arial",14,'bold'),
                    fg='#fff', bg='#f50', relief="raised",pady=2,padx=5)   
    mod_lab8_1 = tk.Label(window, font=("Arial",14,'bold'),
                    width=9,relief="raised",pady=2, padx=5,bg="gray72",fg="blue")
    mod_lab9 = tk.Label(window,text="  會        計  ",font=("Arial",14,'bold'),
                    fg='#fff', bg='#f50', relief="raised",pady=2,padx=5)   
    mod_lab9_1 = tk.Label(window, font=("Arial",14,'bold'),
                    width=9,relief="raised",pady=2, padx=5,bg="gray72",fg="blue")  
    mod_lab10 = tk.Label(window,text="  註 記  ",font=("Arial",14,'bold'),
                    fg='#fff', bg='#f50', relief="raised",pady=2,padx=5)  

    mod_lab11 = tk.Label(window,text="列印前、請再次確認輸入資料是否正確 !",
                    font=("Arial",12,'bold'),fg='#f00',relief='groove', padx=5,pady=2) 
    label_copyright = tk.Label(window,text="Copyright © 2022 Puyuma Workshop ",
                    font=("Arial",8,'bold'),bg="skyblue" ,padx=1,pady=1) 

    def validate(P):
        if str.isdigit(P) or P == '':
            return True
        else:
            return False
    a3 = tk.StringVar()
    a4 = tk.StringVar()
    val = tk.StringVar()
    a6 = tk.StringVar()
    a10 = tk.StringVar()
    a3.set("")    
    a4.set("")
    val.set("")    
    a6.set("")
    a10.set("")
    def alter(event):        
        with open("捐款記錄.txt", mode="r+",encoding="UTF-8") as f:
            for line in f.readlines():
                mod_line = line.split(',') 
                if entry1.get() == mod_line[0]:
                    if mod_line[9] == "作廢":
                        tk.messagebox.showinfo('提醒訊息', '本張收據已作廢、請重新輸入!')  
                        entry1.delete(0,'end')
                        break
                    mod_lab2_1.config(text=(mod_line[1]))
                    a3.set(mod_line[2])    
                    a4.set(mod_line[3])
                    val.set(mod_line[4])    
                    a6.set(mod_line[5])  
                    mod_lab7_1.config(text=(mod_line[6]))
                    mod_lab8_1.config(text=(mod_line[7]))
                    mod_lab9_1.config(text=(mod_line[8]))
                    #a10.set(mod_line[9])
                    break
            else:
                tk.messagebox.showerror('錯誤訊息', '使用者不存在')
                
    entry1=tk.Entry(font=('Arial',14,'bold'),width=10,border=1,bd=5)
    entry3=tk.Entry(textvariable=a3,font=('Arial',14,'bold'),width=35,border=1,bd=5)
    entry4=tk.Entry(textvariable=a4,font=('Arial',14,'bold'),width=35,border=1,bd=5)
    vcmd = (window.register(validate), '%P') #限制只能輸入數字
    entry6=tk.Entry(textvariable=a6,validate='key', validatecommand=vcmd,font=('Arial',14,'bold'),
                    width=10,border=1,bd=5)
    entry10=tk.Entry(textvariable=a10,font=('Arial',14,'bold'),justify="center",width=6,border=1,bd=5)

    def clickbton():
        file_data = "" 
        with open("捐款記錄.txt", mode="r",encoding="UTF-8") as f:
            for line in f:
                if entry1.get() == line.split(",")[0]:
                    line = str(entry1.get())+","+str(line.split(",")[1])+ ","+ str(entry3.get())+ ","+ str(entry4.get())+","+str(entry5.get())+","+str(entry6.get())+","+str(mod_lab7_1.cget("text"))+","+str(mod_lab8_1.cget("text"))+","+str(mod_lab9_1.cget("text"))+","+str(entry10.get())+",\n"
                    file_data += line
                else:    
                    file_data += line    
        with open("捐款記錄.txt", mode="w",encoding="UTF-8") as f:
            f.write(file_data)
        tk.messagebox.showinfo('提醒訊息', '記錄修改成功')

        wb = load_workbook("收據範本1210.xlsx")
        ws = wb.active
        # 存根聯
        ws["f14"].value = "NO. "+str(收據編號產生器.new_ser_No)
        ws["c4"].value = str(int(str(datetime.today())[0:4])-1911)+"年"+str(datetime.today())[5:7]+"月"+str(datetime.today())[8:10]+"日"
        ws["b5"].value = entry3.get()
        ws["b7"].value = entry4.get()
        ws["c8"].value = entry5.get()
        ws["c9"].value = converter.m_convert(entry6.get())
        ws["b10"].value = mod_lab7_1.cget("text")
        ws["f10"].value = mod_lab8_1.cget("text")
        ws["d10"].value = mod_lab9_1.cget("text")
        # 收執聯
        ws["f33"].value = "NO. "+str(收據編號產生器.new_ser_No)
        ws["c23"].value = str(int(str(datetime.today())[0:4])-1911)+"年"+str(datetime.today())[5:7]+"月"+str(datetime.today())[8:10]+"日"
        ws["b24"].value = entry3.get()
        ws["b26"].value = entry4.get()
        ws["c27"].value = entry5.get()
        ws["c28"].value = converter.m_convert(entry6.get())
        ws["b29"].value = mod_lab7_1.cget("text")
        ws["f29"].value = mod_lab8_1.cget("text")
        ws["d29"].value = mod_lab9_1.cget("text")
        # wb.save("中研福德宮"+str(收據編號產生器.new_ser_No)+".xlsx")
        wb.save("收據範本1210.xlsx")
        #printing(wb) ###### 輸出列表機 ######
        #shutil.copy('捐款記錄.txt','e:\.record_backup') # 資料備份到e:\
        entry1.delete(0,'end')    
        entry3.delete(0,'end')
        entry4.delete(0,'end')    
        entry5.delete(0,'end')
        entry6.delete(0,'end')
        mod_lab7_1.config(text="")
        mod_lab8_1.config(text="")
        mod_lab9_1.config(text="")
        entry10.delete(0,'end')
    def ending():
        window.destroy()
        menu()

    entry1.bind('<Return>',alter)
    mod_btn = tk.Button(window,text="儲存 與 列印",font=('Arial',14,'bold'),
                    padx=2,pady=2,activeforeground='#f00', command=clickbton)        
    ending_btn = tk.Button(window,text="離  開",font=('Arial',12,'bold'),
                    fg="white",bg="blue",activeforeground='#f00',command=ending)
    #####################################################
    def enable():
        entry5.config(state="normal")
    def disable():
        entry5.config(state="disable")
    val = tk.StringVar()  # 建立文字變數
    radio_btn1 = tk.Radiobutton(text='喜添捐獻',font=('Arial',12),bg="skyblue",
                                variable=val, value='喜添捐獻',command=disable)
    radio_btn1.place(x=200, y=235)
    radio_btn1.select()
    radio_btn2 = tk.Radiobutton(text='入會費',font=('Arial',12),bg="skyblue",
                                variable=val, value='入會費',command=disable)
    radio_btn2.place(x=305, y=235)
    radio_btn3 = tk.Radiobutton(text='年度會費',font=('Arial',12),bg="skyblue",
                                variable=val, value='年度會費',command=enable)
    radio_btn3.place(x=400, y=235)
    radio_btn4 = tk.Radiobutton(text='其他',font=('Arial',12),bg="skyblue",
                                variable=val, value='其他',command=enable)
    radio_btn4.place(x=500, y=235)
    entry5=tk.Entry(textvariable=val, font=('Arial',14,'bold'),width=15,border=1,bd=5,state="disable")
    ######################################################
    label_head.place(x=200,y=5)
    mod_lab1.place(x=340, y=65)
    entry1.place(x=472, y=62)
    mod_lab2.place(x=70, y=65)
    mod_lab2_1.place(x=200, y=65)
    mod_lab3.place(x=70, y=107)
    entry3.place(x=200, y=105)
    mod_lab4.place(x=70, y=150)
    entry4.place(x=200, y=147)
    mod_lab5.place(x=70, y=191)
    entry5.place(x=200, y=189)
    mod_lab6.place(x=70, y=270)
    entry6.place(x=200, y=267)
    mod_lab7.place(x=70, y=310)
    mod_lab7_1.place(x=200, y=310)
    mod_lab8.place(x=340, y=310)
    mod_lab8_1.place(x=470, y=310)
    mod_lab9.place(x=340, y=270)
    mod_lab9_1.place(x=470, y=270)
    mod_lab10.place(x=430, y=191)
    entry10.place(x=520, y=190)
    mod_lab11.place(x=180, y=350)
    mod_btn.place(x=260, y=390)
    ending_btn.place(x=40, y=400)
    label_copyright.place(x=430, y=430)
    window.mainloop()

if __name__ == "__main__":
    mod_module()    